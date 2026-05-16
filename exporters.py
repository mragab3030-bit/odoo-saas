import io
import logging
import os
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import arabic_reshaper
from bidi.algorithm import get_display

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


logger = logging.getLogger(__name__)

BRAND_BLUE = colors.HexColor('#2563eb')
BRAND_DARK = colors.HexColor('#1e293b')
LIGHT_GREY = colors.HexColor('#f8fafc')
MID_GREY = colors.HexColor('#e2e8f0')


# ---------------------------------------------------------------------------
# Unicode / Arabic font registration
# ---------------------------------------------------------------------------

_FONTS_DIR = os.path.join(os.path.dirname(__file__), 'static', 'fonts')

# Latin / general font — used for non-Arabic cells and the body text.
_LATIN_CANDIDATES = [
    (os.path.join(_FONTS_DIR, 'DejaVuSans.ttf'),
     os.path.join(_FONTS_DIR, 'DejaVuSans-Bold.ttf')),
    ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
     '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
    ('/usr/share/fonts/dejavu/DejaVuSans.ttf',
     '/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf'),
]

# Dedicated Arabic font — applied to cells that contain Arabic glyphs.
_ARABIC_CANDIDATES = [
    (os.path.join(_FONTS_DIR, 'Amiri-Regular.ttf'),
     os.path.join(_FONTS_DIR, 'Amiri-Bold.ttf')),
    ('/usr/share/fonts/truetype/amiri/Amiri-Regular.ttf',
     '/usr/share/fonts/truetype/amiri/Amiri-Bold.ttf'),
]

UNICODE_FONT = 'Helvetica'
UNICODE_FONT_BOLD = 'Helvetica-Bold'
ARABIC_FONT = UNICODE_FONT
ARABIC_FONT_BOLD = UNICODE_FONT_BOLD


def _register_pair(candidates, regular_name, bold_name, fallback):
    for regular, bold in candidates:
        if not os.path.exists(regular):
            continue
        try:
            pdfmetrics.registerFont(TTFont(regular_name, regular))
            reg = regular_name
            if os.path.exists(bold):
                pdfmetrics.registerFont(TTFont(bold_name, bold))
                bld = bold_name
            else:
                bld = reg
            logger.info("PDF font registered (%s): %s", regular_name, regular)
            return reg, bld
        except Exception as e:
            logger.warning("Failed to register font %s: %s", regular, e)
    logger.warning("No font found for %s — using %s.", regular_name, fallback)
    return fallback, fallback


def _register_fonts():
    global UNICODE_FONT, UNICODE_FONT_BOLD, ARABIC_FONT, ARABIC_FONT_BOLD
    UNICODE_FONT, UNICODE_FONT_BOLD = _register_pair(
        _LATIN_CANDIDATES, 'AppFont', 'AppFont-Bold', 'Helvetica')
    ARABIC_FONT, ARABIC_FONT_BOLD = _register_pair(
        _ARABIC_CANDIDATES, 'AppArabic', 'AppArabic-Bold', UNICODE_FONT)


_register_fonts()


def _has_arabic(text: str) -> bool:
    return any('؀' <= ch <= 'ۿ' or 'ݐ' <= ch <= 'ݿ'
               or 'ﭐ' <= ch <= '﷿' or 'ﹰ' <= ch <= '﻿'
               for ch in text)


def _shape_arabic(text: str) -> str:
    try:
        return get_display(arabic_reshaper.reshape(text))
    except Exception:
        return text


def _esc(text: str) -> str:
    return text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def export_pdf(title: str, headers: list, rows: list,
               subtitle: str = '', col_widths: list = None) -> io.BytesIO:
    buf = io.BytesIO()
    page_size = landscape(A4) if len(headers) > 6 else A4
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title', parent=styles['Title'],
        fontName=UNICODE_FONT_BOLD, fontSize=16,
        textColor=BRAND_DARK, spaceAfter=4
    )
    sub_style = ParagraphStyle(
        'Sub', parent=styles['Normal'],
        fontName=UNICODE_FONT, fontSize=9,
        textColor=colors.HexColor('#64748b'), spaceAfter=12
    )
    cell_style_ltr = ParagraphStyle(
        'Cell', parent=styles['Normal'],
        fontName=UNICODE_FONT, fontSize=8,
        textColor=BRAND_DARK, alignment=TA_LEFT,
        leading=10,
    )
    cell_style_rtl = ParagraphStyle(
        'CellRTL', parent=cell_style_ltr,
        fontName=ARABIC_FONT, fontSize=10,
        alignment=TA_RIGHT, leading=13,
    )
    header_style = ParagraphStyle(
        'Header', parent=styles['Normal'],
        fontName=UNICODE_FONT_BOLD, fontSize=9,
        textColor=colors.white, alignment=TA_CENTER,
        leading=11,
    )

    def make_cell(value, header_cell=False):
        text = '' if value is None else str(value)
        if _has_arabic(text):
            text = _shape_arabic(text)
            return Paragraph(_esc(text), cell_style_rtl)
        return Paragraph(_esc(text), header_style if header_cell else cell_style_ltr)

    elements = [
        Paragraph(title, title_style),
        Paragraph(
            subtitle or f'Generated on {datetime.now().strftime("%Y-%m-%d %H:%M")} | {len(rows)} records',
            sub_style
        ),
        Spacer(1, 0.3 * cm),
    ]

    table_data = (
        [[make_cell(h, header_cell=True) for h in headers]]
        + [[make_cell(c) for c in row] for row in rows]
    )

    col_count = len(headers)
    page_width = page_size[0]
    available_width = page_width - 2.4 * cm  # matches left+right margins above

    if col_widths and len(col_widths) == col_count:
        total = float(sum(col_widths)) or 1.0
        widths = [w / total * available_width for w in col_widths]
    else:
        widths = [available_width / col_count] * col_count

    table = Table(table_data, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BRAND_BLUE),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), UNICODE_FONT_BOLD),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
        ('FONTNAME', (0, 1), (-1, -1), UNICODE_FONT),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, MID_GREY),
        ('LINEBELOW', (0, 0), (-1, 0), 1, BRAND_BLUE),
    ]))

    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf


REV_GREEN  = colors.HexColor('#16a34a')
EXP_RED    = colors.HexColor('#dc2626')
NET_BLUE   = colors.HexColor('#2563eb')
NET_RED    = colors.HexColor('#dc2626')
SUBTOTAL_BG = colors.HexColor('#f1f5f9')


def _fmt_amount(value, currency='', unit='units'):
    """Format a money number with optional K / M suffix.

    unit:
      'units' → 1,234,567.89
      'k'     → 1,234.6K   (one decimal, scaled by 1,000)
      'm'     → 1.23M      (two decimals, scaled by 1,000,000)
    """
    try:
        n = float(value or 0)
    except (TypeError, ValueError):
        n = 0.0
    u = (unit or 'units').lower()
    if u == 'k':
        s = '{:,.1f}K'.format(n / 1000.0)
    elif u == 'm':
        s = '{:,.2f}M'.format(n / 1_000_000.0)
    else:
        s = '{:,.2f}'.format(n)
    return ('%s %s' % (currency, s)).strip()


def export_pnl_pdf(company_name: str, cost_center_name: str,
                   period_label: str, currency: str,
                   pnl: dict, unit: str = 'units') -> io.BytesIO:
    """Render a one-page Cost Center P&L Statement.

    pnl shape:
      revenue_lines / expense_lines: [{'name', 'code', 'amount'}, ...]
      total_revenue / total_expenses / net: float
      margin_display: str (e.g. '23.5%' or '—')
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=1.5 * cm, leftMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'Title', parent=styles['Title'],
        fontName=UNICODE_FONT_BOLD, fontSize=18,
        textColor=BRAND_DARK, alignment=TA_CENTER, spaceAfter=6,
    )
    sub_style = ParagraphStyle(
        'Sub', parent=styles['Normal'],
        fontName=UNICODE_FONT, fontSize=10,
        textColor=colors.HexColor('#64748b'),
        alignment=TA_CENTER, spaceAfter=14,
    )
    section_label = ParagraphStyle(
        'SectionLabel', parent=styles['Normal'],
        fontName=UNICODE_FONT_BOLD, fontSize=11,
        textColor=BRAND_DARK, spaceAfter=4,
    )

    def cell(text, bold=False, align='left', color=BRAND_DARK):
        style = ParagraphStyle(
            'C', parent=styles['Normal'],
            fontName=UNICODE_FONT_BOLD if bold else UNICODE_FONT,
            fontSize=10, textColor=color, leading=13,
            alignment={'left': TA_LEFT, 'right': TA_RIGHT,
                       'center': TA_CENTER}[align],
        )
        text = '' if text is None else str(text)
        if _has_arabic(text):
            text = _shape_arabic(text)
        return Paragraph(_esc(text), style)

    elements = [
        Paragraph('COST CENTER P&amp;L STATEMENT', title_style),
    ]
    meta_parts = []
    if company_name:
        meta_parts.append(company_name)
    if cost_center_name:
        meta_parts.append('Cost Center: %s' % cost_center_name)
    if period_label:
        meta_parts.append('Period: %s' % period_label)
    meta_parts.append('Generated %s' %
                      datetime.now().strftime('%Y-%m-%d %H:%M'))
    is_compare = (pnl.get('compare_mode') or 'none') in ('yoy', 'prev')
    compare_label = pnl.get('compare_label') or ''

    if is_compare and compare_label:
        elements.append(Paragraph('  |  '.join(meta_parts), sub_style))
        cmp_style = ParagraphStyle(
            'Cmp', parent=styles['Normal'],
            fontName=UNICODE_FONT, fontSize=9,
            textColor=colors.HexColor('#64748b'),
            alignment=TA_CENTER, spaceAfter=10,
        )
        elements.append(Paragraph(
            'Comparing with: <b>%s</b>' % _esc(compare_label), cmp_style))
    else:
        elements.append(Paragraph('  |  '.join(meta_parts), sub_style))

    page_width = A4[0]
    available_width = page_width - 3.0 * cm
    if is_compare:
        # 5 cols: name (40%) / current (16%) / previous (16%) / change (16%) / chg% (12%)
        section_col_widths = [
            available_width * 0.40,
            available_width * 0.16,
            available_width * 0.16,
            available_width * 0.16,
            available_width * 0.12,
        ]
    else:
        # 3 cols: name (60%) / % (12%) / amount (28%)
        section_col_widths = [
            available_width * 0.60,
            available_width * 0.12,
            available_width * 0.28,
        ]
    NUM_COLS = len(section_col_widths)
    pct_neutral = colors.HexColor('#9ca3af')
    SUBGROUP_BG = colors.HexColor('#f8fafc')
    UP_COLOR    = colors.HexColor('#16a34a')
    DOWN_COLOR  = colors.HexColor('#dc2626')

    def _change_color(diff, kind):
        """kind: 'rev' | 'exp' | 'oth' | 'net'. Net mirrors revenue
        direction (up == good); expenses are inverted."""
        try:
            d = float(diff or 0)
        except (TypeError, ValueError):
            d = 0.0
        if d == 0:
            return pct_neutral
        if kind == 'exp':
            return DOWN_COLOR if d > 0 else UP_COLOR
        return UP_COLOR if d > 0 else DOWN_COLOR

    def _change_text(diff):
        try:
            d = float(diff or 0)
        except (TypeError, ValueError):
            d = 0.0
        if d == 0:
            arrow, sign = '=', ''
        elif d > 0:
            arrow, sign = '▲', '+'
        else:
            arrow, sign = '▼', '-'
        return '%s %s%s' % (arrow, sign, _fmt_amount(abs(d), currency, unit))

    def _line_row(ln, kind, bold=False, label_override=None,
                  amount_override=None, color_override=None):
        """Build one PDF table row for a line OR a subtotal-style entry."""
        amt = amount_override if amount_override is not None else (ln.get('amount') or 0)
        amount_color = color_override or BRAND_DARK
        if is_compare:
            prev_val = ln.get('previous_amount')
            prev_disp = (_fmt_amount(prev_val, currency, unit)
                         if (prev_val is not None and prev_val != 0) or prev_val == 0
                         else '—')
            diff = ln.get('change') or 0
            diff_pct = ln.get('change_pct_display') or '—'
            change_color = _change_color(diff, kind)
            return [
                cell('   ' + (label_override or
                              (('[%s] ' % ln.get('code')) if ln.get('code') else '') +
                              (ln.get('name') or 'Unassigned')),
                     bold=bold),
                cell(_fmt_amount(amt, currency, unit), align='right',
                     bold=bold, color=amount_color),
                cell(prev_disp, align='right', color=pct_neutral),
                cell(_change_text(diff), align='right', bold=bold,
                     color=change_color),
                cell(diff_pct, align='right', bold=bold, color=change_color),
            ]
        # Non-compare layout
        return [
            cell('   ' + (label_override or
                          (('[%s] ' % ln.get('code')) if ln.get('code') else '') +
                          (ln.get('name') or 'Unassigned')),
                 bold=bold),
            cell(ln.get('pct_display') or '—', align='right',
                 bold=bold, color=pct_neutral),
            cell(_fmt_amount(amt, currency, unit), align='right',
                 bold=bold, color=amount_color),
        ]

    def _subtotal_row(label, total, pct_display, kind, color, group=None):
        """Section / sub-group subtotal row."""
        if is_compare:
            prev = (group or {}).get('previous_subtotal') if group else None
            if prev is None and kind == 'rev':
                prev = pnl.get('previous_total_revenue')
            if prev is None and kind == 'exp':
                prev = pnl.get('previous_total_expenses')
            if prev is None and kind == 'oth':
                prev = pnl.get('previous_total_other')
            diff = ((group or {}).get('change') if group else None)
            diff_pct = ((group or {}).get('change_pct_display') if group else None)
            if group is None and kind == 'rev':
                diff = pnl.get('total_revenue_change')
                diff_pct = pnl.get('total_revenue_change_pct_display')
            if group is None and kind == 'exp':
                diff = pnl.get('total_expenses_change')
                diff_pct = pnl.get('total_expenses_change_pct_display')
            if group is None and kind == 'oth':
                diff = pnl.get('total_other_change')
                diff_pct = pnl.get('total_other_change_pct_display')
            change_color = _change_color(diff or 0, kind)
            return [
                cell(label, bold=True),
                cell(_fmt_amount(total, currency, unit), bold=True,
                     align='right', color=color),
                cell(_fmt_amount(prev, currency, unit) if (prev is not None and prev != 0) or prev == 0 else '—',
                     align='right', color=pct_neutral),
                cell(_change_text(diff or 0), bold=True, align='right',
                     color=change_color),
                cell(diff_pct or '—', bold=True, align='right',
                     color=change_color),
            ]
        return [
            cell(label, bold=True),
            cell(pct_display or ('100.0%' if total else '—'),
                 bold=True, align='right', color=color),
            cell(_fmt_amount(total, currency, unit), bold=True,
                 align='right', color=color),
        ]

    def section_block(label, color, groups, flat_lines, total, total_label, kind):
        """Render one section (REVENUE / EXPENSES / OTHER). kind drives
        the comparison colour direction."""
        # Section title row spans all columns.
        header_row = [cell(label, bold=True, color=color)]
        header_row.extend(cell('', align='right') for _ in range(NUM_COLS - 1))
        table_data = [header_row]
        subgroup_total_rows = []

        # If compare mode, add a column-headers row right under the title.
        if is_compare:
            hdr = [cell('Account', bold=True, color=colors.HexColor('#64748b'))]
            for txt in ('Current', 'Previous', 'Change', 'Change %'):
                hdr.append(cell(txt, bold=True, align='right',
                                color=colors.HexColor('#64748b')))
            table_data.append(hdr)

        if groups:
            for grp in groups:
                # Sub-group label row (full-width left-aligned label)
                sg_row = [cell('  ' + (grp.get('label') or ''),
                               bold=True, color=color)]
                sg_row.extend(cell('', align='right')
                              for _ in range(NUM_COLS - 1))
                table_data.append(sg_row)
                # Lines
                for ln in grp.get('lines') or []:
                    table_data.append(_line_row(ln, kind))
                # Sub-group subtotal
                sub = grp.get('subtotal') or 0
                table_data.append(_subtotal_row(
                    '    Subtotal — ' + (grp.get('label') or ''),
                    sub, grp.get('pct_display'), kind, color, group=grp))
                subgroup_total_rows.append(len(table_data) - 1)
        else:
            for ln in (flat_lines or []):
                table_data.append(_line_row(ln, kind))
            if not flat_lines:
                no_row = [cell('   (no lines)',
                               color=colors.HexColor('#94a3b8'))]
                for _ in range(NUM_COLS - 1):
                    no_row.append(cell('—', align='right',
                                       color=colors.HexColor('#94a3b8')))
                table_data.append(no_row)

        # Section total row
        table_data.append(_subtotal_row(
            total_label, total, '100.0%' if total else '—',
            kind, color, group=None))

        t = Table(table_data, colWidths=section_col_widths, hAlign='LEFT')
        styles_list = [
            ('FONTNAME', (0, 0), (-1, -1), UNICODE_FONT),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3.5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3.5),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            # Section header row
            ('BACKGROUND', (0, 0), (-1, 0), SUBTOTAL_BG),
            ('LINEBELOW', (0, 0), (-1, 0), 0.7, color),
            # Section subtotal row
            ('BACKGROUND', (0, -1), (-1, -1), SUBTOTAL_BG),
            ('LINEABOVE', (0, -1), (-1, -1), 0.9, color),
            ('LINEBELOW', (0, -1), (-1, -1), 0.9, color),
        ]
        if is_compare:
            # Column-headers row under the section title.
            styles_list.append(
                ('LINEBELOW', (0, 1), (-1, 1), 0.3,
                 colors.HexColor('#cbd5e1')))
        for r in subgroup_total_rows:
            styles_list.append(('BACKGROUND', (0, r), (-1, r), SUBGROUP_BG))
            styles_list.append(('LINEABOVE', (0, r), (-1, r), 0.4, color))
        t.setStyle(TableStyle(styles_list))
        return t

    elements.append(section_block(
        'REVENUE', REV_GREEN,
        pnl.get('revenue_groups') or [],
        pnl.get('revenue_lines') or [],
        pnl.get('total_revenue') or 0, 'Total Revenue', 'rev'))
    elements.append(Spacer(1, 0.4 * cm))
    elements.append(section_block(
        'EXPENSES', EXP_RED,
        pnl.get('expense_groups') or [],
        pnl.get('expense_lines') or [],
        pnl.get('total_expenses') or 0, 'Total Expenses', 'exp'))
    elements.append(Spacer(1, 0.6 * cm))

    net_value  = pnl.get('net') or 0
    net_color  = NET_BLUE if net_value >= 0 else NET_RED
    net_label  = 'NET PROFIT' if net_value >= 0 else 'NET LOSS'

    if is_compare:
        prev_net = pnl.get('previous_net') or 0
        net_diff = pnl.get('net_change') or 0
        net_diff_pct = pnl.get('net_change_pct_display') or '—'
        net_change_color = _change_color(net_diff, 'net')
        # Summary block uses the same 5-col width so amount columns
        # align with the sections above.
        summary_data = [
            [cell(net_label, bold=True, color=net_color),
             cell(_fmt_amount(net_value, currency, unit),
                  bold=True, align='right', color=net_color),
             cell(_fmt_amount(prev_net, currency, unit) if (prev_net or prev_net == 0) else '—',
                  align='right', color=pct_neutral),
             cell(_change_text(net_diff), bold=True, align='right',
                  color=net_change_color),
             cell(net_diff_pct, bold=True, align='right',
                  color=net_change_color)],
            [cell('Margin %', bold=True),
             cell(pnl.get('margin_display') or '—', bold=True,
                  align='right', color=net_color),
             cell('', align='right'),
             cell('', align='right'),
             cell('', align='right')],
        ]
    else:
        summary_data = [
            [cell(net_label, bold=True, color=net_color),
             cell('', align='right'),
             cell(_fmt_amount(net_value, currency, unit), bold=True, align='right',
                  color=net_color)],
            [cell('Margin %', bold=True),
             cell('', align='right'),
             cell(pnl.get('margin_display') or '—', bold=True, align='right',
                  color=net_color)],
        ]
    summary = Table(summary_data, colWidths=section_col_widths)
    summary.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), UNICODE_FONT_BOLD),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, 1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, 0), (-1, -1), SUBTOTAL_BG),
        ('LINEABOVE', (0, 0), (-1, 0), 1.2, net_color),
        ('LINEBELOW', (0, -1), (-1, -1), 1.2, net_color),
    ]))
    elements.append(summary)

    # Balance-sheet movements live at the bottom so they don't visually
    # compete with the Net result.
    other_lines = pnl.get('other_lines') or []
    if other_lines:
        elements.append(Spacer(1, 0.6 * cm))
        elements.append(section_block(
            'OTHER MOVEMENTS (non-P&amp;L)',
            colors.HexColor('#64748b'),
            None,  # other section has no sub-grouping
            other_lines,
            pnl.get('total_other') or 0,
            'Total Other', 'oth'))

    doc.build(elements)
    buf.seek(0)
    return buf


def export_pnl_excel(company_name: str, cost_center_name: str,
                     period_label: str, currency: str,
                     pnl: dict, unit: str = 'units') -> io.BytesIO:
    """Render a Cost Center P&L Statement as a single-sheet workbook.

    When `pnl['compare_mode']` is 'yoy' or 'prev', the layout switches to
    five columns: Account | Current | Previous | Change | Change % —
    each line's change cell is colour-coded per section direction."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cost Center P&L'[:31]

    is_compare = (pnl.get('compare_mode') or 'none') in ('yoy', 'prev')
    num_cols = 5 if is_compare else 3
    last_col_letter = openpyxl.utils.get_column_letter(num_cols)
    AMT_COL = 2 if is_compare else 3   # current-amount column index
    PREV_COL = 3                        # only used in compare
    CHANGE_COL = 4
    CHANGE_PCT_COL = 5

    u = (unit or 'units').lower()
    if u == 'k':
        money_fmt = '#,##0.0,"K"'
    elif u == 'm':
        money_fmt = '#,##0.00,,"M"'
    else:
        money_fmt = '#,##0.00'
    pct_fmt   = '0.0"%"'
    change_fmt = '"▲ +"#,##0.00;"▼ -"#,##0.00;0.00'
    change_fmt_k = '"▲ +"#,##0.0,"K";"▼ -"#,##0.0,"K";"0.0K"'
    change_fmt_m = '"▲ +"#,##0.00,,"M";"▼ -"#,##0.00,,"M";"0.00M"'
    if u == 'k':
        change_fmt = change_fmt_k
    elif u == 'm':
        change_fmt = change_fmt_m
    change_pct_fmt = '"+"0.0"%";"-"0.0"%";"0.0%"'

    pct_color_rev_font = Font(color='16A34A', size=10)
    pct_color_exp_font = Font(color='DC2626', size=10)
    pct_color_neutral  = Font(color='9CA3AF', size=10)
    pct_bold_rev       = Font(bold=True, color='16A34A', size=10)
    pct_bold_exp       = Font(bold=True, color='DC2626', size=10)
    pct_bold_neutral   = Font(bold=True, color='9CA3AF', size=10)
    prev_font          = Font(color='9CA3AF', size=10)

    bold = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14, color='1E293B')
    meta_font  = Font(color='64748B', size=9, italic=True)
    rev_font   = Font(bold=True, color='16A34A', size=11)
    exp_font   = Font(bold=True, color='DC2626', size=11)
    net_font_pos = Font(bold=True, color='2563EB', size=12)
    net_font_neg = Font(bold=True, color='DC2626', size=12)

    header_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9',
                              fill_type='solid')
    thin = Side(style='thin', color='E2E8F0')
    box  = Border(top=thin, bottom=thin, left=thin, right=thin)

    subgroup_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC',
                                fill_type='solid')

    def _change_font(diff, kind, bold_=True):
        """Pick a colour for the Change / Change % cell based on the
        section direction. For expenses, increase = bad (red)."""
        try:
            d = float(diff or 0)
        except (TypeError, ValueError):
            d = 0.0
        if d == 0:
            colour = '9CA3AF'
        elif kind == 'exp':
            colour = 'DC2626' if d > 0 else '16A34A'
        else:
            colour = '16A34A' if d > 0 else 'DC2626'
        return Font(bold=bold_, color=colour, size=10)

    # Title + meta (span all columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws['A1']
    cell.value = 'COST CENTER P&L STATEMENT'
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 26

    meta_lines = []
    if company_name:     meta_lines.append('Company: ' + company_name)
    if cost_center_name: meta_lines.append('Cost Center: ' + cost_center_name)
    if period_label:     meta_lines.append('Period: ' + period_label)
    if is_compare and pnl.get('compare_label'):
        meta_lines.append('Comparing with: ' + pnl['compare_label'])
    meta_lines.append('Generated: ' + datetime.now().strftime('%Y-%m-%d %H:%M'))
    row = 2
    for line in meta_lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        c = ws.cell(row=row, column=1, value=line)
        c.font = meta_font
        c.alignment = Alignment(horizontal='center')
        row += 1
    row += 1  # blank spacer

    def _write_line(ln, pct_line_font, kind):
        nonlocal row
        label_text = ln.get('name') or 'Unassigned'
        code = ln.get('code')
        if code:
            label_text = '[%s] %s' % (code, label_text)
        ws.cell(row=row, column=1, value='   ' + label_text)
        if is_compare:
            cur = ws.cell(row=row, column=AMT_COL, value=ln.get('amount') or 0)
            cur.number_format = money_fmt
            cur.alignment = Alignment(horizontal='right')
            prev_val = ln.get('previous_amount')
            prev_c = ws.cell(row=row, column=PREV_COL,
                             value=(prev_val if (prev_val is not None) else None))
            if prev_val is None:
                prev_c.value = '—'
            elif prev_val == 0:
                prev_c.value = 0
                prev_c.number_format = money_fmt
            else:
                prev_c.number_format = money_fmt
            prev_c.font = prev_font
            prev_c.alignment = Alignment(horizontal='right')
            diff = ln.get('change') or 0
            d_cell = ws.cell(row=row, column=CHANGE_COL, value=diff)
            d_cell.number_format = change_fmt
            d_cell.alignment = Alignment(horizontal='right')
            d_cell.font = _change_font(diff, kind, bold_=False)
            p_pct = ln.get('change_pct')
            p_cell = ws.cell(row=row, column=CHANGE_PCT_COL)
            if p_pct is None:
                p_cell.value = '—'
            else:
                p_cell.value = p_pct
                p_cell.number_format = change_pct_fmt
            p_cell.font = _change_font(diff, kind, bold_=False)
            p_cell.alignment = Alignment(horizontal='right')
        else:
            pct_val = ln.get('pct')
            pct_cell = ws.cell(row=row, column=2, value=None)
            if pct_val is not None and (pct_val != 0 or (
                    ln.get('pct_display') and ln['pct_display'] != '—')):
                pct_cell.value = pct_val
                pct_cell.number_format = pct_fmt
            else:
                pct_cell.value = ln.get('pct_display') or '—'
            pct_cell.font = pct_line_font
            pct_cell.alignment = Alignment(horizontal='right')
            amt_cell = ws.cell(row=row, column=3, value=ln.get('amount') or 0)
            amt_cell.number_format = money_fmt
            amt_cell.alignment = Alignment(horizontal='right')
        row += 1

    def write_section(label, groups, flat_lines, total_label, total,
                      header_font, color, pct_line_font, pct_bold_font,
                      kind, section_prev_total=None,
                      section_change=None, section_change_pct=None):
        """Write one section. In compare mode, accepts the section's
        previous total + change/pct so the section subtotal row can
        render its diff cells."""
        nonlocal row
        # Section title row (full-width merge)
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=num_cols)
        c = ws.cell(row=row, column=1, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='left', indent=0)
        c.border = box
        row += 1

        # Column-headers row right under the title — compare mode only.
        if is_compare:
            hdr_cells = [
                (1, 'Account'),
                (AMT_COL, 'Current'),
                (PREV_COL, 'Previous'),
                (CHANGE_COL, 'Change'),
                (CHANGE_PCT_COL, 'Change %'),
            ]
            for col, txt in hdr_cells:
                h = ws.cell(row=row, column=col, value=txt)
                h.font = Font(bold=True, color='64748B', size=9)
                h.alignment = Alignment(
                    horizontal='left' if col == 1 else 'right')
            row += 1

        if groups:
            for grp in groups:
                # Sub-group label (full-width)
                ws.merge_cells(start_row=row, start_column=1,
                               end_row=row, end_column=num_cols)
                sg = ws.cell(row=row, column=1,
                             value='  ' + (grp.get('label') or ''))
                sg.font = Font(bold=True, color=color, size=10)
                sg.fill = subgroup_fill
                row += 1
                for ln in grp.get('lines') or []:
                    _write_line(ln, pct_line_font, kind)
                # Sub-group subtotal row
                sub = grp.get('subtotal') or 0
                lbl = ws.cell(row=row, column=1,
                              value='    Subtotal — ' + (grp.get('label') or ''))
                lbl.font = Font(bold=True, color='1E293B', size=10)
                lbl.fill = subgroup_fill
                lbl.border = box
                if is_compare:
                    sub_cur = ws.cell(row=row, column=AMT_COL, value=sub)
                    sub_cur.font = Font(bold=True, color=color, size=10)
                    sub_cur.fill = subgroup_fill
                    sub_cur.border = box
                    sub_cur.number_format = money_fmt
                    sub_cur.alignment = Alignment(horizontal='right')
                    p_sub = grp.get('previous_subtotal')
                    p_cell = ws.cell(row=row, column=PREV_COL)
                    if p_sub is None:
                        p_cell.value = '—'
                    else:
                        p_cell.value = p_sub
                        p_cell.number_format = money_fmt
                    p_cell.font = prev_font
                    p_cell.fill = subgroup_fill
                    p_cell.border = box
                    p_cell.alignment = Alignment(horizontal='right')
                    diff = grp.get('change') or 0
                    d_cell = ws.cell(row=row, column=CHANGE_COL, value=diff)
                    d_cell.number_format = change_fmt
                    d_cell.font = _change_font(diff, kind)
                    d_cell.fill = subgroup_fill
                    d_cell.border = box
                    d_cell.alignment = Alignment(horizontal='right')
                    p_pct = grp.get('change_pct')
                    pct_change_cell = ws.cell(row=row, column=CHANGE_PCT_COL)
                    if p_pct is None:
                        pct_change_cell.value = '—'
                    else:
                        pct_change_cell.value = p_pct
                        pct_change_cell.number_format = change_pct_fmt
                    pct_change_cell.font = _change_font(diff, kind)
                    pct_change_cell.fill = subgroup_fill
                    pct_change_cell.border = box
                    pct_change_cell.alignment = Alignment(horizontal='right')
                else:
                    p_val = grp.get('pct')
                    pct_c = ws.cell(row=row, column=2)
                    if p_val is not None and (
                            p_val != 0 or (grp.get('pct_display') and
                                           grp['pct_display'] != '—')):
                        pct_c.value = p_val
                        pct_c.number_format = pct_fmt
                    else:
                        pct_c.value = grp.get('pct_display') or '—'
                    pct_c.font = pct_bold_font
                    pct_c.fill = subgroup_fill
                    pct_c.border = box
                    pct_c.alignment = Alignment(horizontal='right')
                    amt_c = ws.cell(row=row, column=3, value=sub)
                    amt_c.font = Font(bold=True, color=color, size=10)
                    amt_c.fill = subgroup_fill
                    amt_c.border = box
                    amt_c.number_format = money_fmt
                    amt_c.alignment = Alignment(horizontal='right')
                row += 1
        else:
            if not (flat_lines or []):
                empty = ws.cell(row=row, column=1, value='   (no lines)')
                empty.font = Font(color='94A3B8', italic=True)
                row += 1
            for ln in flat_lines or []:
                _write_line(ln, pct_line_font, kind)

        # Section total row
        total_cell_label = ws.cell(row=row, column=1, value=total_label)
        total_cell_label.font = bold
        total_cell_label.fill = header_fill
        total_cell_label.border = box
        if is_compare:
            tcur = ws.cell(row=row, column=AMT_COL, value=total or 0)
            tcur.font = Font(bold=True, color=color, size=11)
            tcur.fill = header_fill
            tcur.border = box
            tcur.number_format = money_fmt
            tcur.alignment = Alignment(horizontal='right')
            tprev = ws.cell(row=row, column=PREV_COL)
            if section_prev_total is None:
                tprev.value = '—'
            else:
                tprev.value = section_prev_total
                tprev.number_format = money_fmt
            tprev.font = prev_font
            tprev.fill = header_fill
            tprev.border = box
            tprev.alignment = Alignment(horizontal='right')
            tdiff = section_change or 0
            td_cell = ws.cell(row=row, column=CHANGE_COL, value=tdiff)
            td_cell.number_format = change_fmt
            td_cell.font = _change_font(tdiff, kind)
            td_cell.fill = header_fill
            td_cell.border = box
            td_cell.alignment = Alignment(horizontal='right')
            tdpct_cell = ws.cell(row=row, column=CHANGE_PCT_COL)
            if section_change_pct is None:
                tdpct_cell.value = '—'
            else:
                tdpct_cell.value = section_change_pct
                tdpct_cell.number_format = change_pct_fmt
            tdpct_cell.font = _change_font(tdiff, kind)
            tdpct_cell.fill = header_fill
            tdpct_cell.border = box
            tdpct_cell.alignment = Alignment(horizontal='right')
        else:
            pct_sub = ws.cell(row=row, column=2)
            if total:
                pct_sub.value = 100.0
                pct_sub.number_format = pct_fmt
            else:
                pct_sub.value = '—'
            pct_sub.font = pct_bold_font
            pct_sub.fill = header_fill
            pct_sub.border = box
            pct_sub.alignment = Alignment(horizontal='right')
            total_cell = ws.cell(row=row, column=3, value=total or 0)
            total_cell.font = Font(bold=True, color=color, size=11)
            total_cell.fill = header_fill
            total_cell.border = box
            total_cell.number_format = money_fmt
            total_cell.alignment = Alignment(horizontal='right')
        row += 2  # spacer

    write_section(
        'REVENUE',
        pnl.get('revenue_groups') or [],
        pnl.get('revenue_lines') or [],
        'Total Revenue', pnl.get('total_revenue') or 0,
        rev_font, '16A34A',
        pct_color_rev_font, pct_bold_rev, 'rev',
        section_prev_total=pnl.get('previous_total_revenue'),
        section_change=pnl.get('total_revenue_change'),
        section_change_pct=pnl.get('total_revenue_change_pct'),
    )
    write_section(
        'EXPENSES',
        pnl.get('expense_groups') or [],
        pnl.get('expense_lines') or [],
        'Total Expenses', pnl.get('total_expenses') or 0,
        exp_font, 'DC2626',
        pct_color_exp_font, pct_bold_exp, 'exp',
        section_prev_total=pnl.get('previous_total_expenses'),
        section_change=pnl.get('total_expenses_change'),
        section_change_pct=pnl.get('total_expenses_change_pct'),
    )
    # ---- Net + margin summary ----
    net_value = pnl.get('net') or 0
    net_color = '2563EB' if net_value >= 0 else 'DC2626'
    net_font  = net_font_pos if net_value >= 0 else net_font_neg
    net_label = 'NET PROFIT' if net_value >= 0 else 'NET LOSS'

    net_label_cell = ws.cell(row=row, column=1, value=net_label)
    net_label_cell.font = net_font
    net_label_cell.fill = header_fill
    net_label_cell.border = box
    if is_compare:
        ncur = ws.cell(row=row, column=AMT_COL, value=net_value)
        ncur.font = net_font; ncur.fill = header_fill; ncur.border = box
        ncur.number_format = money_fmt
        ncur.alignment = Alignment(horizontal='right')
        pnet = pnl.get('previous_net')
        nprev = ws.cell(row=row, column=PREV_COL)
        if pnet is None:
            nprev.value = '—'
        else:
            nprev.value = pnet
            nprev.number_format = money_fmt
        nprev.font = prev_font; nprev.fill = header_fill; nprev.border = box
        nprev.alignment = Alignment(horizontal='right')
        ndiff = pnl.get('net_change') or 0
        ndc = ws.cell(row=row, column=CHANGE_COL, value=ndiff)
        ndc.number_format = change_fmt
        ndc.font = _change_font(ndiff, 'net')
        ndc.fill = header_fill; ndc.border = box
        ndc.alignment = Alignment(horizontal='right')
        npct = pnl.get('net_change_pct')
        npct_cell = ws.cell(row=row, column=CHANGE_PCT_COL)
        if npct is None:
            npct_cell.value = '—'
        else:
            npct_cell.value = npct
            npct_cell.number_format = change_pct_fmt
        npct_cell.font = _change_font(ndiff, 'net')
        npct_cell.fill = header_fill; npct_cell.border = box
        npct_cell.alignment = Alignment(horizontal='right')
    else:
        blank_pct = ws.cell(row=row, column=2, value=None)
        blank_pct.fill = header_fill; blank_pct.border = box
        net_amt_cell = ws.cell(row=row, column=3, value=net_value)
        net_amt_cell.font = net_font
        net_amt_cell.fill = header_fill
        net_amt_cell.border = box
        net_amt_cell.number_format = money_fmt
        net_amt_cell.alignment = Alignment(horizontal='right')
    row += 1

    # Margin % row (current period only — same in both layouts)
    margin_label_cell = ws.cell(row=row, column=1, value='Margin %')
    margin_label_cell.font = bold
    margin_label_cell.fill = header_fill
    margin_label_cell.border = box
    target_col = AMT_COL if is_compare else 3
    for c in range(2, num_cols + 1):
        cc = ws.cell(row=row, column=c, value=None)
        cc.fill = header_fill
        cc.border = box
    margin_value = pnl.get('margin_pct') or 0
    if (pnl.get('total_revenue') or 0) > 0:
        m = ws.cell(row=row, column=target_col, value=margin_value)
        m.number_format = pct_fmt
    else:
        m = ws.cell(row=row, column=target_col,
                    value=pnl.get('margin_display') or '—')
    m.font = Font(bold=True, color=net_color, size=11)
    m.fill = header_fill
    m.border = box
    m.alignment = Alignment(horizontal='right')
    row += 2  # spacer between Net/Margin and the Other section below

    # Balance-sheet movements sit at the bottom so they don't visually
    # compete with the Net result.
    other_lines = pnl.get('other_lines') or []
    if other_lines:
        other_font = Font(bold=True, color='64748B', size=11)
        write_section(
            'OTHER MOVEMENTS (non-P&L)',
            [],  # no sub-grouping for balance-sheet movements
            other_lines,
            'Total Other', pnl.get('total_other') or 0,
            other_font, '64748B',
            pct_color_neutral, pct_bold_neutral, 'oth',
            section_prev_total=pnl.get('previous_total_other'),
            section_change=pnl.get('total_other_change'),
            section_change_pct=pnl.get('total_other_change_pct'),
        )

    # Column widths
    if is_compare:
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
    else:
        ws.column_dimensions['A'].width = 55
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 22

    if currency:
        # money cells use the formatting; symbol is shown in the on-screen
        # currency dropdown in Excel. We keep this hook for future use.
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_excel(title: str, headers: list, rows: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    title_font = Font(bold=True, size=14, color='1E293B')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    meta_cell = ws.cell(
        row=2, column=1,
        value=f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}  |  {len(rows)} records'
    )
    meta_cell.font = Font(color='64748B', size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.row_dimensions[2].height = 18

    header_row = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = Border(
            bottom=Side(style='medium', color='1D4ED8')
        )
    ws.row_dimensions[header_row].height = 22

    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    border = Border(
        bottom=Side(style='thin', color='E2E8F0'),
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
    )

    for row_idx, row in enumerate(rows, header_row + 1):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 18

    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ''))
             for r in range(header_row, header_row + len(rows) + 1)),
            default=8
        )
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = f'A{header_row}:{openpyxl.utils.get_column_letter(len(headers))}{header_row}'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
